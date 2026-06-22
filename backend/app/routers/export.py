import io
from datetime import date

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session, joinedload

from app.core.database import get_db
from app.core.deps import get_current_user
from app.models.intervention import Intervention
from app.models.user import User

router = APIRouter(prefix="/export", tags=["Export"])

# Couleurs reprises du fichier Excel d'origine (statuts)
_STATUT_COLORS = {
    "Terminée": "C6EFCE",
    "En cours": "FFEB9C",
    "Planifiée": "BDD7EE",
    "Annulée": "FFC7CE",
    "Reportée": "F2DCDB",
}

_HEADERS = [
    "#",
    "Date planif.",
    "Client",
    "Technicien",
    "Support",
    "Type d'interv.",
    "Équipement",
    "Nb véh.",
    "Priorité",
    "Statut",
    "Date réalisation",
    "Durée (h)",
    "Commentaires",
    "Preuve / BL",
]


@router.get("/interventions.xlsx")
def export_interventions(
    date_debut: date | None = Query(None),
    date_fin: date | None = Query(None),
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user),
):
    query = db.query(Intervention).options(
        joinedload(Intervention.client),
        joinedload(Intervention.technicien),
        joinedload(Intervention.support),
        joinedload(Intervention.type_intervention),
        joinedload(Intervention.equipement),
        joinedload(Intervention.priorite),
        joinedload(Intervention.statut),
    )
    if date_debut:
        query = query.filter(Intervention.date_planification >= date_debut)
    if date_fin:
        query = query.filter(Intervention.date_planification <= date_fin)

    interventions = query.order_by(Intervention.date_planification).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Suivi des Interventions"

    ws.append(["📋  TABLEAU DE SUIVI DES INTERVENTIONS - COMAFRIQUE TECHNOLOGIES"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(_HEADERS))
    ws["A1"].font = Font(bold=True, size=14)

    ws.append([f"Export généré le {date.today().strftime('%d/%m/%Y')}"])
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(_HEADERS))
    ws.append([])

    header_row_idx = 4
    ws.append(_HEADERS)
    for cell in ws[header_row_idx]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color="305496")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for idx, item in enumerate(interventions, start=1):
        row = [
            idx,
            item.date_planification.strftime("%d/%m/%Y")
            if item.date_planification
            else "",
            item.client.nom if item.client else "",
            item.technicien.nom if item.technicien else "",
            item.support.nom if item.support else "",
            item.type_intervention.libelle if item.type_intervention else "",
            item.equipement.libelle if item.equipement else "",
            item.nb_vehicules or "",
            item.priorite.libelle if item.priorite else "",
            item.statut.libelle if item.statut else "",
            item.date_realisation.strftime("%d/%m/%Y")
            if item.date_realisation
            else "",
            item.duree_heures or "",
            item.commentaires or "",
            item.preuve_bl_url or "",
        ]
        ws.append(row)

        statut_libelle = item.statut.libelle if item.statut else None
        color = _STATUT_COLORS.get(statut_libelle)
        if color:
            statut_cell = ws.cell(row=ws.max_row, column=10)
            statut_cell.fill = PatternFill("solid", start_color=color)

    for col_idx in range(1, len(_HEADERS) + 1):
        column_letter = get_column_letter(col_idx)
        max_length = 10
        for row_idx in range(header_row_idx, ws.max_row + 1):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"Planning_Interventions_Comafrique_{date.today().isoformat()}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
